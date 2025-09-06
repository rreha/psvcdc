import os
import time
import sys
import subprocess
import shutil
import openpyxl

def pause_cli():
    if sys.platform.startswith("win32"):
        os.system("pause")
    else:
        pause = input("Press ENTER to continue...")

def print_exit(txt):
    print(txt)
    print(">    Exiting.")
    pause_cli()
    sys.exit()

def clear_screen():
    if sys.platform.startswith("win32"):
        os.system("cls")
    else:
        os.system("clear")

def pkg_dec():
    if sys.platform.startswith("win32"):
        return "./bin/win/pkg_dec.exe"
    elif sys.platform.startswith("linux"):
        return "./bin/ubuntu64/pkg_dec"
    elif sys.platform.startswith("darwin"):
        return "./bin/macarm64/pkg_dec"

def psvpfsparser():
    if sys.platform.startswith("win32"):
        return "./bin/win/psvpfsparser.exe"
    elif sys.platform.startswith("linux"):
        return "./bin/ubuntu64/psvpfsparser"
    elif sys.platform.startswith("darwin"):
        return "./bin/macarm64/psvpfsparser"

def get_input():
    if len(sys.argv) > 1:
        args_list = []
        for arg in sys.argv[1:]:
            args_list.append(arg)
        print(f">    Got {len(args_list)} input(s).")

    else:
        print_exit("/!\\    Please specify a PS Vita .pkg file or a PS Vita folder (addcont/app/patch).")
    return args_list

def extract_pkg(pkg):
    os.makedirs("./tmp", exist_ok=True)
    subprocess.run([pkg_dec(), "--make-dirs=ux", pkg, "./tmp/"], stdout=subprocess.PIPE, stderr=subprocess.STDOUT, stdin=subprocess.DEVNULL)
    print(">    PKG extracted.")

def detect_content(i):
    if os.path.basename(i) == "addcont":
        print(">    DLC (addcont) detected.")
        return True
    elif os.path.basename(i) == "app":
        print(">    Game (app) detected.")
    elif os.path.basename(i) == "patch":
        print(">    Game Update/Patch (patch) detected.")
    else:
        print_exit("/!\\    Couldn't detect content type. Please make sure to specify addcont/app/patch folder instead of game folder (PCSXXXXXX) itself.")

def get_content_id(i):
    content_id_list = []
    for folder in os.listdir(i):
        path = os.path.join(i, folder)
        if os.path.isdir(path) and folder.startswith("PCS"):
            content_id_list.append(folder)
    return content_id_list

def get_dlc_id(i):
    dlc_id_list = []
    for folder in os.listdir(i):
        path = os.path.join(i, folder)
        if os.path.isdir(path) and folder.startswith("PCS"):
            for dlc in os.listdir(path):
                dlc_path = os.path.join(path, dlc)
                if os.path.isdir(dlc_path):
                    dlc_id_list.append(dlc)
    return dlc_id_list

def get_zrif(content_id, is_dlc, dlc_id=None):
    found = False
    page = "GAMES"
    if is_dlc:
        page = "DLC"
    search = content_id
    missing_values = (None, '-', 'MISSING')
    zrif = None
    db = openpyxl.load_workbook("db.xlsx")
    sheet = db[page]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == search:
                if is_dlc:
                    psn_id = sheet.cell(row=cell.row, column=cell.column + 6)
                    if dlc_id in psn_id.value:
                        result = sheet.cell(row=cell.row, column=cell.column + 5)
                        found = True
                        zrif = result.value
                        break
                else:
                    result = sheet.cell(row=cell.row, column=cell.column + 7)
                    zrif = result.value
                    found = True
                    break
            if found:
                break

    if zrif not in missing_values:
        print(f'>    Found zRIF key ({zrif}).')
        return zrif

    else:
        print_exit("/!\\    Couldn't find zRIF.")

def decrypt_pfs(i, content_id, zrif, dlc_id_list=None, dlc_id=None):
    content = os.path.basename(i)
    os.makedirs("./Decrypted", exist_ok=True)
    print(">    Decrypting PFS.")
    if content == "addcont":
        ret = subprocess.run([psvpfsparser(), "-i", f"{i}/{content_id}/{dlc_id}", "-o", f"./Decrypted/{content}/{content_id}/{dlc_id}", "-z", zrif, "-f", "cma.henkaku.xyz"], stdout=subprocess.PIPE, stderr=subprocess.PIPE, stdin=subprocess.DEVNULL)

    else:        
        ret = subprocess.run([psvpfsparser(), "-i", f"{i}/{content_id}", "-o", f"./Decrypted/{content}/{content_id}", "-z", zrif, "-f", "cma.henkaku.xyz"], stdout=subprocess.PIPE, stderr=subprocess.PIPE, stdin=subprocess.DEVNULL)

    if "invalid" in str(ret):
        print_exit('/!\\    psvpfsparser returned "header signature is invalid". This DLC CANNOT be decrypted using zRIF.')
            
    elif "failed to find unicv.db file or icv.db folder" in str(ret):
        print_exit('/!\\    psvpfsparser returned "failed to find unicv.db file or icv.db folder". This DLC CANNOT be decrypted using zRIF.')
    
    if content == "addcont":
        print(f">    Saved decrypted content to Decrypted/{content}/{content_id}/{dlc_id}.")

    else:
        print(f">    Saved decrypted content to Decrypted/{content}/{content_id}.")

def decrypt_eboot(zrif):
    os.makedirs("./tmp", exist_ok=True)
    for root, dirs, files in os.walk("Decrypted"):
        for folder in dirs:
            folder_path = os.path.join(root, folder)
            eboot_path = os.path.join(folder_path, "eboot.bin")
            if os.path.exists(eboot_path) and not os.path.exists(folder_path + "/eboot_decrypted.bin"):
                print(">    Decrypting eboot.bin")
                subprocess.run([sys.executable, "./util/zrif2rif.py", zrif, "./tmp/work.bin"])
                subprocess.run([sys.executable, "./util/self2elf.py", "-i", eboot_path, "-o", os.path.join(folder_path, "eboot_decrypted.bin"), "-k", "./tmp/work.bin"], stdout=subprocess.PIPE, stderr=subprocess.STDOUT, stdin=subprocess.DEVNULL)
                if os.path.isfile(os.path.join(folder_path, "eboot_decrypted.bin")):
                    if os.path.exists("./util/__pycache__"):
                        shutil.rmtree("./util/__pycache__")
                    print(f">    Saved decrypted eboot.bin (eboot_decrypted.bin) to {folder_path}/eboot_decrypted.bin.")
                else:
                    print_exit("/!\\    eboot.bin couldn't be decrypted. Please check if you have all of the required modules installed.")