import os
import time
import sys
import subprocess
import shutil
import openpyxl

if sys.platform.startswith("linux") or sys.platform.startswith("darwin") == True:
    print("ERROR!!! Sorry, your OS is not supported.")
    pause = input("Press ENTER to continue...")
    sys.exit()

def get_cont_id():
    global cont_id
    cont_id = ""
    flag = False
    for folder in os.listdir("./"):
        if flag == False and folder.startswith("PCS"):
            cont_id = str(folder)
            print(">    Found content ID:", cont_id)
            flag = not flag


def detect_type():
    global content
    global db_sheet
    global cont_id
    global dlc_id
    cont_id = ""
    for folder in os.listdir("./"):
        if folder.startswith("PCS"):
            shutil.rmtree(f"./{folder}")
            
    if os.path.isdir(input_f):
        print(">    Folder detected")
        for folder in os.listdir("./"):
            if folder == "app":
                print(">    Detected Game (app)")
                for f in os.listdir("./app/"):
                    src = os.path.join("./app", f)
                    dest = os.path.join("./", f)
                    shutil.move(src, dest)
                os.rmdir("./app")
                content = "app"
                db_sheet = "GAMES"
                get_cont_id()
                return
            
            elif folder == "patch":
                print(">    Detected Game Update (patch)")
                for f in os.listdir("./patch/"):
                    src = os.path.join("./patch", f)
                    dest = os.path.join("./", f)
                    shutil.move(src, dest)
                os.rmdir("./patch")
                content = "patch"
                db_sheet = "GAMES"
                get_cont_id()
                return
            
            elif folder == "addcont":
                print(">    Detected DLC (addcont)")
                for f in os.listdir("./addcont/"):
                    src = os.path.join("./addcont", f)
                    dest = os.path.join("./", f)
                    shutil.move(src, dest)
                os.rmdir("./addcont")
                content = "addcont"
                db_sheet = "DLC"
                print(">    Organized files")
                get_cont_id()
                cont_id_folder = f"./{cont_id}"
            
                for f in os.listdir("./"):
                    if os.path.isdir(cont_id_folder):
                        for folder in os.listdir(cont_id_folder):
                            dlc_id = folder
                            src = os.path.join(cont_id_folder, folder)
                            dest = os.path.join("./", folder)
                            shutil.move(src, dest)
                    else:
                        print("ERROR!!! Error while organizing DLC. Please run the program again.")
                        os.system("pause")
                        sys.exit()
                    
                print(">    Found DLC ID:", dlc_id)
                return
        
    elif os.path.isfile(input_f):
        print(">    File detected")
        pkgjob()
        
    else:
        print("ERROR!!! Couldn't detect if input is a file or a folder. Maybe it doesn't exist? Please run the program again.")
        os.system("pause")
        sys.exit()
    
def organize_cont():
    global content
    global db_sheet
    global dlc_id
    for folder in os.listdir("./"):
        if folder == "app":
            print(">    Detected Game (app)")
            for f in os.listdir("./app/"):
                src = os.path.join("./app", f)
                dest = os.path.join("./", f)
                shutil.move(src, dest)
            os.rmdir("./app")
            content = "app"
            db_sheet = "GAMES"
            print(">    Organized files")
            get_cont_id()
            return
            
        elif folder == "patch":
            print(">    Detected Game Update (patch)")
            for f in os.listdir("./patch/"):
                src = os.path.join("./patch", f)
                dest = os.path.join("./", f)
                shutil.move(src, dest)
            os.rmdir("./patch")
            content = "patch"
            db_sheet = "GAMES"
            print(">    Organized files")
            get_cont_id()
            return
            
        elif folder == "addcont":
            print(">    Detected DLC (addcont)")
            for f in os.listdir("./addcont/"):
                src = os.path.join("./addcont", f)
                dest = os.path.join("./", f)
                shutil.move(src, dest)
            os.rmdir("./addcont")
            content = "addcont"
            db_sheet = "DLC"
            print(">    Organized files")
            get_cont_id()
            cont_id_folder = "./" + cont_id
            
            for f in os.listdir("./"):
                if os.path.isdir(cont_id_folder):
                    for folder in os.listdir(cont_id_folder):
                        dlc_id = folder
                        src = os.path.join(cont_id_folder, folder)
                        dest = os.path.join("./", folder)
                        shutil.move(src, dest)
                else:
                    print("ERROR!!! Error while organizing DLC. Please run the program again.")
                    os.system("pause")
                    sys.exit()
                    
            print(">    Found DLC ID:", dlc_id)
            return
    return
            
def pkgjob():
    global cont_id
    if os.path.splitext(input_f)[1] == ".pkg":
        print(">    File is a .pkg")
        print(">    Running pkg2zip")
        subprocess.run(["./bin/pkg2zip.exe", "-x", input_f], stdout=subprocess.PIPE, stderr=subprocess.STDOUT, stdin=subprocess.DEVNULL)
        organize_cont()
        
    else:
        print("ERROR!!! Please make sure that the specified file is a .pkg file.")
        os.system("pause")
        sys.exit()
    return
    
def get_zRIF():
    global db_sheet
    print(f">    Finding zRIF key for {cont_id} (might take a while)")
    if cont_id == "" or None:
        print("ERROR!!! Couldn't detect content ID. pkg2zip might not be able to extract the pkg. Try running the program again.")
        os.system("pause")
        sys.exit()
        
    search = cont_id
    zrif = None
    
    wb = openpyxl.load_workbook("db.xlsx")
    sheet = wb[db_sheet]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == search:
                if content == "addcont":
                    next_cell = sheet.cell(row=cell.row, column=cell.column + 5)
                    
                else:
                    next_cell = sheet.cell(row=cell.row, column=cell.column + 7)
                zrif = next_cell.value
                break
        if zrif is not None:
            break

    if zrif != None or '-' or 'MISSING': #for some reason it still doesn't give an error when zrif is - or MISSING???
        print(f'>    Found zRIF ({zrif})')
    else:
        print("ERROR!!! Couldn't find zRIF!")
        if input_f.endswith('.pkg'):
            shutil.rmtree("./{}".format(cont_id))
        os.system("PAUSE")
        sys.exit()
        
        
    if os.path.exists("./DECRYPTED"):
        print(">    Found DECRYPTED folder")
    else:
        print(">    Creating DECRYPTED folder")
        os.mkdir("./DECRYPTED")
    print(">    Running psvpfsparser")
    if content == "addcont":
        ret = subprocess.run(["./bin/psvpfsparser/psvpfsparser.exe", "-i", f"./{dlc_id}", "-o", f"./DECRYPTED/{dlc_id}", "-z", zrif, "-f", "cma.henkaku.xyz"], stdout=subprocess.PIPE, stderr=subprocess.PIPE, stdin=subprocess.DEVNULL)
        ret_str = str(ret)
        subprocess.run(["./bin/psvpfsparser/psvpfsparser.exe", "-i", f"./{dlc_id}", "-o", f"./DECRYPTED/{dlc_id}", "-z", zrif, "-f", "cma.henkaku.xyz"], stdout=subprocess.PIPE, stderr=subprocess.PIPE, stdin=subprocess.DEVNULL)
        if "invalid" in ret_str:
            print('ERROR!!! psvpfsparser returned "header signature is invalid". This DLC CANNOT be decrypted...')
            if input_f.endswith('.pkg') and os.path.exists(f"./{cont_id}") and os.path.exists(f"./{dlc_id}"):
                shutil.rmtree(f"./{cont_id}")
                shutil.rmtree(f"./{dlc_id}")
                shutil.rmtree("./DECRYPTED")
            os.system("pause")
            sys.exit()
            
        elif "failed to find unicv.db file or icv.db folder" in ret_str:
            print('ERROR!!! psvpfsparser returned "failed to find unicv.db file or icv.db folder". This DLC CANNOT be decrypted...')
            if input_f.endswith('.pkg') and os.path.exists(f"./{cont_id}") and os.path.exists(f"./{dlc_id}"):
                shutil.rmtree(f"./{cont_id}")
                shutil.rmtree(f"./{dlc_id}")
                shutil.rmtree("./DECRYPTED")
            os.system("pause")
            sys.exit()
            
        if not os.path.exists(f"./DECRYPTED/{dlc_id}"):
            print("ERROR!!! An error occurred while checking if content was decrypted successfully. Maybe zRIF is invalid? Please run the program again.")
            if input_f.endswith('.pkg') and os.path.exists(f"./{cont_id}") and os.path.exists(f"./{dlc_id}"):
                shutil.rmtree(f"./{cont_id}")
                shutil.rmtree(f"./{dlc_id}")
                shutil.rmtree("./DECRYPTED")
            os.system("pause")
            sys.exit()
            
        src = os.path.join("./", dlc_id)
        dest = os.path.join("./", cont_id)
        shutil.move(src, dest)
        os.mkdir(f"./DECRYPTED/{content}")
        os.mkdir(f"./DECRYPTED/{content}/{cont_id}")
        src = os.path.join(f"./DECRYPTED/{dlc_id}")
        dest = os.path.join(f"./DECRYPTED/{content}/{cont_id}")
        shutil.move(src, dest)

    else:
        if not os.path.exists(f"./DECRYPTED/{content}"):
            os.mkdir(f"./DECRYPTED/{content}")
        subprocess.run(["./bin/psvpfsparser/psvpfsparser.exe", "-i", f"./{cont_id}", "-o", f"./DECRYPTED/{content}/{cont_id}","-z", zrif, "-f", "cma.henkaku.xyz"], stdout=subprocess.PIPE, stderr=subprocess.STDOUT, stdin=subprocess.DEVNULL)
        os.mkdir(f"{content}")
        src = os.path.join(f"./{cont_id}")
        dest = os.path.join(f"./{content}/{cont_id}")
        shutil.move(src, dest)
        
    print(f">    Saved decrypted content to /DECRYPTED/{content}/{cont_id}.")
    
    rm_encrypted = input(">    Delete encrypted content? (Y/N): ")
    if rm_encrypted.lower() == "y":
        shutil.rmtree(f"./{content}")
        print(">    Deleted encrypted content")
    else:
        print(">    Keeping encrypted content")
    if content != "addcont":
        dec_eboot = input(">    Decrypt eboot.bin? (Y/N): ")
        if dec_eboot.lower() == "y":
            print(">    Converting zRIF to work.bin")
            subprocess.run(["python", "zrif2rif.py", zrif])
            print(">    Decrypting eboot.bin")
            subprocess.run(["python", "self2elf.py", "-i", f"DECRYPTED/{content}/{cont_id}/eboot.bin", "-o", f"DECRYPTED/{content}/{cont_id}/eboot_decrypted.bin", "-k", "work.bin"], stdout=subprocess.PIPE, stderr=subprocess.STDOUT, stdin=subprocess.DEVNULL)
            shutil.rmtree("./__pycache__")
            os.remove("work.bin")
            print(f">    Decrypted eboot.bin (DECRYPTED/{content}/{cont_id}/eboot_decrypted.bin)")
            return
        
        else:
            return
        
        return

    
# main

print("PS Vita Content DeCryptor v1.1 ~ https://github.com/rreha/psvcdc\n")
try:
    global input_f
    input_f = sys.argv[1] 
    print(">    Input:", input_f)
    
except IndexError:
    print("Please specify a PS Vita .pkg file or a PS Vita folder")
    os.system("pause")
    sys.exit()
 
detect_type()
get_zRIF()

print(">    Exiting...")
os.system("pause")
sys.exit()